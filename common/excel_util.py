import os

import openpyxl


class ExcelUtil:
    #获得项目路径
    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]


    def read_excel(self,sheetname="login"):
        #加载excel
        wb=openpyxl.load_workbook(self.get_object_path()+"data/login_data.xlsx")
        sheet=wb[sheetname]
        #获得excel行数和列数
        all_list=[]
        for rows in range(2,sheet.max_row+1):
            temp_list=[]
            for cols in range(1,sheet.max_column+1):
                temp_list.append(sheet.cell(rows,cols).value)
            all_list.append(temp_list)
        return all_list




